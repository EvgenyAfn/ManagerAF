import tkinter as tk
import tkinter.font as tkfont
import customtkinter as ctk
import openpyxl
import subprocess
import bcrypt
import pyzipper
import os
import random
import string
from tkinter import filedialog
from tkinter.scrolledtext import ScrolledText
from tktooltip import ToolTip
from PIL import Image, ImageTk
from tkinter import ttk


row_index = 0  # Объявление переменной row_index вне функций
excel_filename = ""  # Define excel_filename at the module level
login_mail = "af_mail"  # Ввести логин для авторизации отправки сообщений(From)
host = "smtp.nikoil.ru"
basedir = os.path.dirname(__file__)


def select_excel_file():
    global excel_filename  # Use the global keyword to modify the module-level variable
    excel_filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not excel_filename:
        return
    log_action(f"✅Выбран Excel файл: {excel_filename}")
    update_chose_exc_label()


def update_chose_exc_label():
    if excel_filename:
        ftt = tkfont.Font(family='Segoe UI Emoji', size=8)
        chose_exc.config(text=f"📝 {excel_filename}", fg="green", justify="left", font=ftt)


def search_user():
    login = login_entry.get()
    if not excel_filename or not login:
        log_error("❗Выберите Excel файл и введите логин❗")
        return

    try:
        workbook = openpyxl.load_workbook(excel_filename)
        sheet = workbook.active

        user_row = None
        row_number = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_number += 1
            if row[14] == login:
                user_row = list(row)
                break

        if user_row is None:
            log_error("Пользователь с указанным логином не найден.")
        else:
            password, hash_password = user_row[15], user_row[16]
            if password is None:
                password = generate_password()
                user_row[15] = password
                log_action(f"Сгенерирован пароль: {password}")
            else:
                log_warning("Ячейка с паролем уже заполнена.")

            if hash_password is None:
                hash_password = generate_hash(password)
                user_row[16] = hash_password
                log_action(f"Сгенерирован и записан хеш пароля: {hash_password}")
            else:
                log_warning("Ячейка с хешем пароля уже заполнена.")

            # Преобразуем кортеж user_row в список и обновляем значения
            for i, value in enumerate(user_row, start=1):
                sheet.cell(row=row_number + 1, column=i, value=value)

        workbook.save(excel_filename)
    except Exception as e:
        log_error(f"Ошибка при обработке файла: {str(e)}")


def generate_password():
    # Генерация пароля из 10 случайных символов (буквы и цифры)
    import string
    import random
    characters = string.ascii_letters + string.digits
    password = ''.join(random.choice(characters) for _ in range(10))
    return password


def generate_hash(password):
    # Генерация хеша пароля с использованием bcrypt версии 3.2.0
    salt = bcrypt.gensalt(rounds=8, prefix=b'2a')
    hash_password = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hash_password.decode('utf-8')


# def generate_hash(password):
#     # Генерация хеша пароля с использованием bcrypt 4.0.+
#     salt = bcrypt.gensalt(rounds=8)
#     hash_password = bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')
#     return hash_password


def save_password_to_txt():
    login = login_entry.get()  # Получаем введенный логин из поля ввода

    if not login:
        log_error("Введите логин, чтобы сохранить пароль в .txt")
        return

    try:
        workbook = openpyxl.load_workbook(excel_filename)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[14] == login:
                password = row[15]  # Получаем пароль из столбца 'P'

                if password is not None:
                    password = str(password)
                    with open(f"./txt/{login}.txt", "w") as txt_file:
                        txt_file.write(password)
                    log_action(f"🔐Пароль сохранен в файле: {login}.txt")
                else:
                    log_warning("⚠️Ячейка с паролем пуста!")
                break
        else:
            log_error("❗Пользователь с указанным логином не найден")

    except Exception as e:
        log_error(f"Ошибка при обработке файла: {str(e)}")


def create_and_encrypt_zip_archive():
    login = login_entry.get()  # Получаем введенный логин из поля ввода
    if not login:
        log_error("Введите логин, чтобы сохранить пароль в .zip")
        return
    txt_file_path = f"./txt/{login}.txt"
    zip_file_path = f"./to_send/{login}.zip"
    password = generate_random_password()

    try:
        if not os.path.isfile(txt_file_path):
            log_error(f"❗Текстового файла для пользователя {login} не существует.")
            return

        workbook = openpyxl.load_workbook(excel_filename)
        sheet = workbook.active
        email = None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[14] == login:
                email = row[7]
                break

        if email:
            log_action(f"🗄️Zip архив создан и сохранен: {login}.zip")
            log_info(f"Пароль от архива: {password} - {email}")
        else:
            log_action(f"🗄️Zip архив создан и сохранен: {login}.zip")
            log_info(f"Пароль от архива: {password} - ⚠️empty Email⚠️")

        with pyzipper.AESZipFile(zip_file_path, 'w', compression=pyzipper.ZIP_LZMA,
                                 encryption=pyzipper.WZ_AES) as zip_file:
            zip_file.setpassword(password.encode())
            zip_file.write(txt_file_path, f"{login}.txt")

    except Exception as e:
        log_error(f"Ошибка при создании архива: {str(e)}")


def generate_random_password():
    # Генерация случайного пароля для ZIP из 5 символов (латинские буквы и цифры)
    characters = string.ascii_letters + string.digits
    password = ''.join(random.choice(characters) for _ in range(5))
    return password


def generate_sql_script(login, hash_password):
    if not login:
        log_error("Введите логин для генерации SQL-скрипта.")
        return None

    if not hash_password:
        log_error("Хеш пароля не найден. Выполните Pass&Hash для этого пользователя.")
        return None

    sql_script = f"insert into WEB_API_Users(login,password,refreshTokenId,RoleId,UserId,weight)\n"
    sql_script += f"SELECT '{login}','{hash_password}',NULL,'2C4A4212-A2EA-42D6-B932-B1720FAFBD53',NEWID(),1;"

    return sql_script


def show_sql_window():
    login = login_entry.get()

    if not login:
        log_error("Введите логин для генерации SQL-скрипта.")
        return

    try:
        workbook = openpyxl.load_workbook(excel_filename)
        sheet = workbook.active

        user_exists = False
        hash_password = None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[14] == login:
                user_exists = True
                hash_password = row[16]
                break

        if not user_exists:
            log_error("Пользователь с указанным логином не найден.")
        elif hash_password is None:
            log_error("Хеш пароля не найден. Выполните Pass&Hash для этого пользователя.")
        else:
            sql_script = generate_sql_script(login, hash_password)
            log_action(f"✔️Ваш SQL-скрипт для пользователя {login} составлен.")

            if sql_script:
                x_offset = root.winfo_x() - 750  # Задайте желаемое смещение слева от родительского окна
                y_offset = root.winfo_y() + 180
                sql_window = tk.Toplevel(root)
                sql_window.title("SQL-скрипт")
                sql_window.geometry(f"+{x_offset}+{y_offset}")
                sql_window.resizable(width=False, height=False)
                sql_text = tk.Text(sql_window, wrap=tk.WORD, width=90, height=8)
                sql_text.insert(tk.END, sql_script)
                sql_text.pack()
                sql_text.config(state=tk.DISABLED)

                def copy_and_close():
                    sql_script_to_copy = sql_text.get("1.0", "end-1c")
                    sql_window.clipboard_clear()
                    sql_window.clipboard_append(sql_script_to_copy)
                    sql_window.update()  # now it stays on the clipboard after the window is closed
                    sql_window.destroy()

                copy_button = tk.Button(sql_window, text="📝Скопировать в буфер и закрыть❌",
                                        cursor="hand2", command=copy_and_close)
                copy_button.pack()
    except Exception as e:
        log_error(f"Ошибка при обработке файла: {str(e)}")


def generate_email_pattern(login):
    if not login:
        log_error("Введите логин для составления шаблона.")
        return None

    mail_pattern = f"Добрый день!\n"
    mail_pattern += (f"Для доступа к веб-клиенту системы Актив-Факторинг Банка Уралсиб используйте ссылку "
                     f"https://factoring.uralsib.ru/.\n")
    mail_pattern += f"Логин - '{login}'\n"
    mail_pattern += f"Пароль от логина в архиве. Пароль для архива будет выслан отдельным письмом.\n"
    mail_pattern += (f"При возникновении проблем с подключением, просьба писать на support_factoring@uralsib.ru, "
                     f"с указанием Наименования контрагента, ИНН, ФИО контактного лица.\n")

    return mail_pattern


def show_mail_window():
    login = login_entry.get()

    if not login:
        log_error("Введите логин для составления шаблона.")
        return

    try:
        workbook = openpyxl.load_workbook(excel_filename)
        sheet = workbook.active

        user_exists = False

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[14] == login:
                user_exists = True
                break

        if not user_exists:
            log_error("Пользователь с указанным логином не найден.")
        else:
            mail_pattern = generate_email_pattern(login)
            log_action(f"📩Ваш шаблон письма для {login} составлен.")

            if mail_pattern:
                x_offset = root.winfo_x() - 750  # Задайте желаемое смещение слева от родительского окна
                y_offset = root.winfo_y() - 20
                mail_window = tk.Toplevel(root)
                mail_window.title("Шаблон для письма")
                mail_window.geometry(f"+{x_offset}+{y_offset}")
                mail_text = tk.Text(mail_window, wrap=tk.WORD, width=90, height=8)
                mail_text.insert(tk.END, mail_pattern)
                mail_text.pack()
                mail_text.config(state=tk.DISABLED)
                mail_window.resizable(width=False, height=False)

                def copy_and_close():
                    mail_pattern_to_copy = mail_text.get("1.0", "end-1c")
                    mail_window.clipboard_clear()
                    mail_window.clipboard_append(mail_pattern_to_copy)
                    mail_window.update()  # now it stays on the clipboard after the window is closed
                    mail_window.destroy()

                copy_button = tk.Button(mail_window, text="📝Скопировать в буфер и закрыть❌",
                                        cursor="hand2", command=copy_and_close)
                copy_button.pack()
                mail_window.mainloop()
    except Exception as e:
        log_error(f"Ошибка при обработке файла: {str(e)}")

class AnimatedButton(tk.Label):
    def __init__(self, parent, click_function):
        super().__init__(parent)

        self.image_list = []
        self.current_frame = 0
        self.play_animation = False
        self.animation_active = False
        self.click_function = click_function

        base_dir = os.path.dirname(__file__)  # Директория, в которой находится скрипт
        sending_dir = os.path.join(base_dir, 'icons', 'sendings')

        # Load images from the folder
        for i in range(1, 92):
            image_name = f"frame-{i:02d}.png"
            image_path = os.path.join(sending_dir, image_name)
            if os.path.exists(image_path):
                image = Image.open(image_path)
                image = image.resize((160, 100))
                self.image_list.append(ImageTk.PhotoImage(image))

        if self.image_list:
            self.configure(image=self.image_list[self.current_frame])
            self.bind("<Button-1>", self.toggle_animation)

    def toggle_animation(self, _event):
        if not self.animation_active:
            if self.play_animation:
                self.play_animation = False
            else:
                self.animation_active = True
                self.current_frame = 0
                self.animate()
                self.click_function()

    def animate(self):
        if self.animation_active and self.current_frame < len(self.image_list):
            self.configure(image=self.image_list[self.current_frame])
            self.current_frame += 1
            self.after(25, self.animate)
        else:
            self.animation_active = False


class SlidePanel(ctk.CTkFrame):
    def __init__(self, parent, start_pos, end_pos):
        super().__init__(master=parent)

        # general attributes
        self.start_pos = start_pos + 0.07
        self.end_pos = end_pos - 0.173
        # self.width = abs(start_pos - end_pos)
        self.width = 0.465
        self.configure(fg_color="#979aaa")

        # animation logic
        self.pos = self.start_pos
        self.in_start_pos = True

        # layout
        self.place(relx=self.start_pos, rely=0.025, relwidth=self.width, relheight=0.75)

    def animate(self, _event=None):
        if self.in_start_pos:
            self.animate_forward()
        else:
            self.animate_backwards()

    def animate_forward(self):
        if self.pos > self.end_pos:
            self.pos -= 0.008
            self.place(relx=self.pos, rely=0.025, relwidth=self.width, relheight=0.75)
            self.after(5, self.animate_forward)
        else:
            self.in_start_pos = False

    def animate_backwards(self, _event=None):
        if self.pos < self.start_pos:
            self.pos += 0.008
            self.place(relx=self.pos, rely=0.025, relwidth=self.width, relheight=0.75)
            self.after(5, self.animate_backwards)
        else:
            self.in_start_pos = True


def open_email_conf_dialog():
    login = login_entry.get()

    if not login:
        log_error("Введите логин для отправки письма")
        return

    try:
        workbook = openpyxl.load_workbook(excel_filename)
        sheet = workbook.active

        user_exists = False
        email = None
        mail_pattern = generate_email_pattern(login)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[14] == login:
                user_exists = True
                email = row[7]
                break

        if not user_exists:
            log_error("Пользователь с указанным логином не найден.")
        elif email is None:
            log_warning("Email пользователя не указан. Невозможно отправить письмо.")

        else:
            mail_conf_window = tk.Toplevel(root)
            x_offset = root.winfo_x() + 510  # Задать желаемое смещение от родительского окна
            y_offset = root.winfo_y() - 20
            mail_conf_window.geometry(f"460x350+{x_offset}+{y_offset}")
            mail_conf_window.configure(bg="#85A1B7")
            mail_conf_window.title("Параметры и отправка Email")
            mail_conf_window.iconbitmap(os.path.join(basedir, "icons", "favicon_gear.ico"))
            mail_conf_window.resizable(width=False, height=False)
            log_warning("ВНИМАНИЕ: этот функционал пока тестируется и может работать некорректно.")

            # Виджеты окна
            from_label = tk.Label(mail_conf_window, text="From:", bg="#85A1B7", fg="#FFFFFF")
            from_label.place(x=10, y=20, width=70, height=25)
            from_entry = tk.Entry(mail_conf_window, bg="#6D7D89", fg="#FFFFFF")
            from_entry.insert(0, "support_factoring@uralsib.ru")
            from_entry.place(x=65, y=20, width=170, height=25)

            to_label = tk.Label(mail_conf_window, text="To Send:", bg="#85A1B7", fg="#FFFFFF")
            to_label.place(x=3, y=60, width=70, height=25)
            to_entry = tk.Entry(mail_conf_window, bg="#6D7D89", fg="#FFFFFF")
            to_entry.insert(0, f"{email}")
            to_entry.place(x=65, y=60, width=170, height=25)

            subj_label = tk.Label(mail_conf_window, text="Subject:", bg="#85A1B7", fg="#FFFFFF")
            subj_label.place(x=3, y=90, width=70, height=25)
            subj_entry = tk.Text(mail_conf_window, bg="#6D7D89", fg="#FFFFFF")
            subj_entry.insert(tk.END, "Доступ к веб-клиенту Актив-Факторинг. Банк Уралсиб")
            subj_entry.configure(font=("Arial", 9))
            subj_entry.place(x=10, y=115, width=225, height=35)

            mess_label = tk.Label(mail_conf_window, text="Message:", bg="#85A1B7", fg="#FFFFFF")
            mess_label.place(x=3, y=155, width=70, height=25)
            mess_entry = tk.Text(mail_conf_window, bg="#6D7D89", fg="#FFFFFF")
            mess_entry.insert(tk.END, f"{mail_pattern}")
            mess_entry.configure(font=("Arial", 9))
            mess_entry.place(x=10, y=180, width=440, height=150)

            def open_to_send_folder():
                folder_path = './to_send'  # Change this to the folder you want to open
                if os.path.exists(folder_path):
                    subprocess.Popen(['explorer', os.path.abspath(folder_path)])

            def label_click(_event):
                if zip_icon_path == os.path.join(basedir, "icons", "arch.png"):
                    open_to_send_folder()

            zip_file_path = f"./to_send/{login}.zip"
            if os.path.exists(zip_file_path):
                zip_icon_path = os.path.join(basedir, "icons", "arch.png")
            else:
                zip_icon_path = os.path.join(basedir, "icons", "error.png")

            zip_label = tk.Label(mail_conf_window, text="Zip:", bg="#85A1B7", fg="#FFFFFF")
            zip_label.place(x=175, y=155, width=40, height=25)

            # Load and resize the icon (archive or error)
            icon_size = (23, 23)  # Set your desired icon size here
            icon_image = Image.open(zip_icon_path)
            icon_image = icon_image.resize(icon_size)  # Resize the icon
            zip_icon = ImageTk.PhotoImage(icon_image)  # Convert the resized image to PhotoImage

            zip_icon_label = tk.Label(mail_conf_window, image=zip_icon, bg="#85A1B7", cursor="hand2")
            zip_icon_label.place(x=210, y=150)
            zip_icon_label.image = zip_icon  # Keep a reference to the image
            # Bind the label to the click event
            zip_icon_label.bind("<Button-1>", label_click)
            ToolTip(zip_icon_label, msg="Открыть папку с ZIP", delay=1)

            def send_email_on_click():
                if not validate_login_and_password():
                    return

                use_tls = check_sec_var.get()

                send_email(from_entry.get(), to_entry.get(),
                           subj_entry.get("1.0", "end-1c"),
                           mess_entry.get("1.0", "end-1c"),
                           port_entry.get(),
                           host_entry.get(),
                           login_auth.get(),
                           pass_auth.get(),
                           use_tls)

            # Кнопочка отправить письмо
            # image_label = Extra(mail_conf_window, send_email_on_click, label_width=160, label_height=100,
            #                     image_width=160, image_height=100, cursor="hand2")
            # image_label.pack()
            # image_label.place(x=250, y=60)

            animated_button = AnimatedButton(mail_conf_window, send_email_on_click)
            animated_button.pack(expand=True)
            animated_button.config(cursor="hand2")
            animated_button.place(x=250, y=60, width=150, height=100)

            ToolTip(animated_button, msg="Отправить сообщение", delay=1)

            # Тут жесть начинается
            wrench_image_path = os.path.join(basedir, "icons", "wrench.png")
            wrench_icon_size = (30, 30)

            wrench_image = Image.open(wrench_image_path)
            wrench_image = wrench_image.resize(wrench_icon_size)
            wrench_image = ImageTk.PhotoImage(wrench_image)

            wrench_label = tk.Label(mail_conf_window, image=wrench_image, text="", bg="#85A1B7", fg="#FFFFFF",
                                    cursor="hand2")
            wrench_label.place(relx=0.95, rely=0.08, anchor='center')

            # Создаем объект SlidePanel
            animated_panel = SlidePanel(mail_conf_window, 1.0, 0.7)

            wrench_label.bind("<Button-1>", lambda event: animated_panel.animate())

            ToolTip(wrench_label, msg="Открыть панель конфигурации", delay=1)

            conf_label = tk.Label(animated_panel, text="Configuration", bg="#979aaa", fg="#FFFFFF")
            conf_label.configure(font=("Segoe UI Emoji", 11, "bold"))
            conf_label.place(x=-15, y=-5, width=145, height=32)

            host_label = tk.Label(animated_panel, text="SMTP host:", bg="#979aaa")
            host_label.configure(font=("Segoe UI Emoji", 11, "bold"), fg="#bb1818")
            host_label.place(x=5, y=25, width=81, height=30)
            host_entry = tk.Entry(animated_panel, bg="#6D7D89", fg="#FFFFFF")
            host_entry.insert(tk.END, f"{host}")
            host_entry.place(x=90, y=30, width=114, height=25)

            port_label = tk.Label(animated_panel, text="Port", bg="#979aaa", fg="#FFFFFF")
            port_label.configure(font=("Segoe UI Emoji", 11))
            port_label.place(x=5, y=60, width=78, height=30)
            port_entry = tk.Entry(animated_panel, bg="#6D7D89", fg="#FFFFFF")
            port_entry.insert(tk.END, "25")
            port_entry.place(x=90, y=65, width=114, height=25)

            style = ttk.Style()
            style.configure("TCheckbutton", background="#979aaa", foreground="#FFFFFF", activeforeground="#FFFFFF",
                            activebackground="#979aaa")
            # Чек-бокс Use Sec
            check_sec_var = tk.IntVar()
            check_sec_var.set(0)
            check_sec = ttk.Checkbutton(animated_panel, variable=check_sec_var, text=" Use Secured Connection",
                                        style="TCheckbutton")
            check_sec.place(x=5, y=95, width=180, height=25)

            # Чек-бокс Use Auth
            check_auth_var = tk.IntVar()
            check_auth_var.set(0)
            use_auth = ttk.Checkbutton(animated_panel, text=" Use Auth:", variable=check_auth_var,
                                       style="TCheckbutton")
            use_auth.place(x=5, y=120, width=81, height=30)

            l_auth_label = tk.Label(animated_panel, text="Login:", bg="#979aaa", fg="#FFFFFF")
            l_auth_label.place(x=5, y=148, width=46, height=30)
            login_auth = tk.Entry(animated_panel, state="disabled", bg="#6D7D89", fg="#FFFFFF")
            login_auth.place(x=50, y=150, width=154, height=25)

            p_auth_label = tk.Label(animated_panel, text="Pass:", bg="#979aaa", fg="#FFFFFF")
            p_auth_label.place(x=5, y=183, width=45, height=30)
            pass_auth = tk.Entry(animated_panel, show='*', state="disabled", bg="#6D7D89", fg="#FFFFFF")
            pass_auth.place(x=50, y=185, width=127, height=25)

            # def on_label_click(_event):
            #     animated_panel.animate_backwards()

            ok_icon_size = (30, 30)
            ok_image_path = os.path.join(basedir, "icons", "ok.png")
            my_image2 = Image.open(ok_image_path)  # Загрузка изображения
            my_image2 = my_image2.resize(ok_icon_size)
            my_image2 = ImageTk.PhotoImage(my_image2)  # Преобразуйте изображение в PhotoImage
            ok_label = tk.Label(animated_panel, image=my_image2, bg="#979aaa", cursor="hand2")
            ok_label.place(x=90, y=220)
            ok_label.image = my_image2  # Установите изображение в атрибут image
            ok_label.bind("<Button-1>", lambda event: animated_panel.animate_backwards())
            ToolTip(ok_label, msg="Применить настройки", delay=1)

            # Скрываем пароль под звездочки
            def toggle_password_visibility():
                if pass_auth.cget('show') == '':
                    pass_auth.config(show='*')
                else:
                    pass_auth.config(show='')

            # Кнопка с глазиком
            show_password_button = tk.Button(animated_panel, text="👁",
                                             command=toggle_password_visibility, state="disabled",
                                             borderwidth=0, bg="#979aaa", fg="#FFFFFF", activeforeground="#FFFFFF",
                                             activebackground="#979aaa")
            show_password_button.configure(font=("Segoe UI Emoji", 11))
            show_password_button.place(x=178, y=185, width=35, height=25)
            ToolTip(show_password_button, msg="Показать пароль", delay=1)

            # Функция для активации/деактивации полей при изменении состояния чекбокса
            def toggle_auth_fields():
                if check_auth_var.get() == 1:
                    login_auth.config(state="normal")
                    login_auth.insert(0, f"{login_mail}")
                    pass_auth.config(state="normal")
                    show_password_button.config(state="normal", cursor="hand2")
                else:
                    pass_auth.delete(0, "end")
                    login_auth.delete(0, "end")
                    login_auth.config(state="disabled")
                    pass_auth.config(state="disabled")
                    show_password_button.config(state="disabled")
                    pass_auth.config(show='*')

            check_auth_var.trace("w", lambda *args: toggle_auth_fields())

            def validate_login_and_password():
                if check_auth_var.get() == 1:  # Checkbox is active
                    login_au = login_auth.get()
                    password = pass_auth.get()
                    if not login_au:
                        log_error("Введите логин.")
                        return False
                    if not password:
                        log_error("Введите пароль.")
                        return False
                return True

            mail_conf_window.mainloop()
    except Exception as e:
        log_error(f"Ошибка: {str(e)}")


def send_email(from_email, to_email, email_subject, mess_content, port_content, host_content, login_au, password,
               use_tls):
    import smtplib
    from email.mime.application import MIMEApplication
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.utils import formatdate
    import ssl

    login = login_entry.get()

    if not login:
        log_error("Введите логин для отправки письма.")
        return

    if not from_email:
        log_error("Введите адрес отправителя.")
        return

    if not to_email:
        log_error("Введите адрес получателя.")
        return

    if not email_subject:
        log_error("Введите тему письма.")
        return

    if not mess_content:
        log_error("Введите текст письма.")
        return

    if not port_content:
        log_error("Введите порт.")
        return

    if not host_content:
        log_error("Введите HOST.")
        return

    try:
        workbook = openpyxl.load_workbook(excel_filename)
        sheet = workbook.active

        user_exists = False
        email = None
        context = None  # Создайте объект контекста для TLS

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[14] == login:
                user_exists = True
                email = row[7]
                break

        if not user_exists:
            log_error("Пользователь с указанным логином не найден.")
        elif email is None:
            log_warning("Email пользователя не указан. Невозможно отправить письмо.")
        else:
            mail_pattern = mess_content

            if mail_pattern:
                msg = MIMEMultipart()
                msg['From'] = from_email
                msg['To'] = to_email
                msg['Date'] = formatdate(localtime=True)
                msg['Subject'] = email_subject
                part2 = MIMEText(mail_pattern, 'plain', "utf-8")
                msg.attach(part2)

                # Прикрепляем сгенерированный архив
                zip_file_path = f"./to_send/{login}.zip"
                if os.path.exists(zip_file_path):
                    with open(zip_file_path, "rb") as file:
                        part = MIMEApplication(
                            file.read(),
                            Name=f"{login}.zip")
                    part['Content-Disposition'] = f'attachment; filename="{login}.zip"'
                    msg.attach(part)
                else:
                    log_error("❗Архив для пользователя не сгенерирован или переименован.")
                    return

                # Устанавливаем соединение с SMTP-сервером и отправляем письмо
                if use_tls:
                    context = ssl.create_default_context()
                    context.options &= ~ssl.OP_SINGLE_ECDH_USE
                    context.options &= ~ssl.OP_NO_COMPRESSION

                with smtplib.SMTP(host_content, port_content) as smtp:
                    if use_tls:
                        smtp.starttls(context=context)
                    smtp.login(login_au, password)
                    smtp.sendmail(msg['From'], msg['To'], msg.as_string().encode('utf-8'))
                    log_action(f"✔️Письмо отправлено на адрес: {to_email}")

    except Exception as e:
        log_error(f"❗Ошибка при отправке письма (проверьте корректность введенных данных): {str(e)}")
        log_warning(f"⚠️Проверьте корректность введенных данных!")


def log_action(message):
    log_text.config(state=tk.NORMAL)
    log_text.insert(tk.END, f"{message}\n")
    log_text.config(state=tk.DISABLED)


def log_error(message):
    log_text.config(state=tk.NORMAL)
    log_text.insert(tk.END, f"Error: {message}\n", "error")
    log_text.config(state=tk.DISABLED)


def log_info(message):
    log_text.config(state=tk.NORMAL)
    log_text.insert(tk.END, f"Info: {message}\n", "info")
    log_text.config(state=tk.DISABLED)


def log_warning(message):
    log_text.config(state=tk.NORMAL)
    log_text.insert(tk.END, f"Warning: {message}\n", "warning")
    log_text.config(state=tk.DISABLED)


def clear_log():
    login_entry.delete(0, tk.END)
    log_text.config(state=tk.NORMAL)
    log_text.delete(1.0, tk.END)
    log_text.config(state=tk.DISABLED)


# Инициализация GUI
root = tk.Tk()
root.title("Manager AF (#0.14)")
width = 485
height = 315
screenwidth = root.winfo_screenwidth()
screenheight = root.winfo_screenheight()
alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
root.geometry(alignstr)
root.resizable(width=False, height=False)
root.iconbitmap(os.path.join(basedir, "icons", "favicon_main.ico"))

# Надпись о выборе Excel
chose_exc = tk.Label(root)
chose_exc["cursor"] = "arrow"
ft = tkfont.Font(family='Segoe UI Emoji', size=10)
chose_exc["font"] = ft
chose_exc["fg"] = "red"
chose_exc["text"] = "Для начала выберите Excel-файл ➡️"
chose_exc["wraplength"] = 355
chose_exc.place(x=0, y=4, width=355, height=45)

# Кнопка "Выберете Excel-файл"
select_excel_button = tk.Button(root)
select_excel_button["cursor"] = "hand2"
ft = tkfont.Font(family='Segoe UI Emoji', size=10)
select_excel_button["font"] = ft
select_excel_button["fg"] = "#000000"
select_excel_button["justify"] = "center"
select_excel_button["text"] = "Выбрать 📂"
excelTip = ToolTip(select_excel_button, msg="Выберите Excel-файл", delay=1)
select_excel_button.place(x=360, y=10, width=100, height=30)
select_excel_button["command"] = select_excel_file

# Логин
login_label = tk.Label(root)
ft = tkfont.Font(family='Segoe UI Emoji', size=10)
login_label["font"] = ft
login_label["fg"] = "#333333"
login_label["justify"] = "right"
login_label["text"] = "Логин:"
login_label.place(x=40, y=80, width=70, height=25)

# Панель ввода логина
login_entry = tk.Entry(root)
login_entry["borderwidth"] = "1px"
ft = tkfont.Font(family='Segoe UI Emoji', size=10)
login_entry["font"] = ft
login_entry["fg"] = "#333333"
login_entry["justify"] = "left"
loginTip = ToolTip(login_entry, msg="Введите имя пользователя", delay=1)
login_entry.place(x=111, y=80, width=160, height=30)

# Кнопка "Pass&Hash"
search_button = tk.Button(root)
search_button["bg"] = "#f0f0f0"
ft = tkfont.Font(family='Segoe UI Emoji', size=10)
search_button["cursor"] = "hand2"
search_button["font"] = ft
search_button["fg"] = "#000000"
search_button["justify"] = "center"
search_button["text"] = "Pass&Hash 🔑"
phTip = ToolTip(search_button, msg="🔑 Сгенерировать пароль/hash пользователю", delay=1)
search_button.place(x=280, y=80, width=100, height=30)
search_button["command"] = search_user

# Текстовое поле для лога
log_text = ScrolledText(root, wrap=tk.WORD, height=10)
log_text.tag_config("error", foreground="red")
log_text.tag_config("info", foreground="blue")
log_text.tag_config("warning", foreground="orange")
ft = tkfont.Font(family='Segoe UI Emoji', size=10)
log_text["bg"] = "#f0f0f0"
log_text["font"] = ft
log_text["state"] = tk.DISABLED
log_text["fg"] = "#333333"
log_text["relief"] = "sunken"
log_text.place(x=90, y=130, width=385, height=161)

# Кнопка TXT
save_txt = tk.Button(root)
save_txt["bg"] = "#f0f0f0"
ft = tkfont.Font(family='Segoe UI Emoji', size=9)
save_txt["font"] = ft
save_txt["fg"] = "#000000"
save_txt["cursor"] = "hand2"
save_txt["justify"] = "center"
save_txt["text"] = "🗒️TXT"
txtTip = ToolTip(save_txt, msg="🗒️Создать .txt-файл с паролем в папку ./txt", delay=1)
save_txt.place(x=10, y=130, width=70, height=25)
save_txt["command"] = save_password_to_txt

# Кнопка ZIP
save_zip = tk.Button(root)
save_zip["bg"] = "#f0f0f0"
ft = tkfont.Font(family='Segoe UI Emoji', size=10)
save_zip["font"] = ft
save_zip["fg"] = "#000000"
save_zip["justify"] = "center"
save_zip["cursor"] = "hand2"
save_zip["text"] = "🗄️ZIP"
ZipTip = ToolTip(save_zip, msg="🗄️Создать Zip-архив с файлом txt в папку ./to_send", delay=1)
save_zip.place(x=10, y=160, width=70, height=25)
save_zip["command"] = create_and_encrypt_zip_archive

# Кнопка ClearLog
cls_log = tk.Button(root)
cls_log["bg"] = "#f0f0f0"
ft = tkfont.Font(family='Segoe UI Emoji', size=10)
cls_log["font"] = ft
cls_log["fg"] = "#000000"
cls_log["justify"] = "center"
cls_log["cursor"] = "hand2"
cls_log["text"] = "🗑️Clear"
LogTip = ToolTip(cls_log, msg="🗑️ Очистить окно лога и поле ввода логина", delay=1)
cls_log.place(x=390, y=80, width=70, height=30)
cls_log["command"] = clear_log

# Кнопка SQL
sql_button = tk.Button(root)
sql_button["bg"] = "#f0f0f0"
ft = tkfont.Font(family='Times', size=10)
sql_button["font"] = ft
sql_button["fg"] = "#000000"
sql_button["justify"] = "center"
sql_button["cursor"] = "hand2"
sql_button["text"] = "♻️SQL"
sqlTip = ToolTip(sql_button, msg="♻️ Сгенерировать SQL-Скрипт", delay=1)
sql_button.place(x=10, y=220, width=70, height=25)
sql_button["command"] = show_sql_window

# Кнопка Email
mail = tk.Button(root)
mail["bg"] = "#f0f0f0"
ft = tkfont.Font(family='Segoe UI Emoji', size=10)
mail["font"] = ft
mail["fg"] = "#000000"
mail["justify"] = "center"
mail["cursor"] = "hand2"
mail["text"] = "📧Sample"
mailTip = ToolTip(mail, msg="Сгенерировать Email пользователю", delay=1)
mail.place(x=10, y=190, width=70, height=25)
mail["command"] = show_mail_window

# Кнопка c отправкой письма
lotus_button = tk.Button(root)
lotus_button["bg"] = "#f0f0f0"
fnt = tkfont.Font(family='Segoe UI Emoji', size=10)
lotus_button["font"] = fnt
lotus_button["fg"] = "#000000"
lotus_button["cursor"] = "hand2"
lotus_button["justify"] = "center"
lotus_button["text"] = "📪 (beta)"
lotusTip = ToolTip(lotus_button, msg="Отправить письмо с конфигурациями", delay=1)
lotus_button.place(x=10, y=260, width=70, height=30)
lotus_button["command"] = open_email_conf_dialog

root.focus_force()
root.mainloop()
